"""Convert Excel spreadsheets into other useful file types."""

import sys
import os
import time
import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import pyfiglet
import json
import xml.etree.ElementTree as elementTree
import threading
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics

# Prevent directory errors
if getattr(sys, 'frozen', False):
    # If running as a bundled executable
    base_path = sys._MEIPASS
else:
    # If running in a normal Python environment
    base_path = os.path.dirname(__file__)

# Imported font 
font_path = os.path.join(base_path, 'DejaVuSans.ttf')

# Register the font
pdfmetrics.registerFont(TTFont('DejaVu', font_path))
# Prepare the ASCII text banner
ascii_banner = pyfiglet.figlet_format("TURASAS", font="colossal")


class Color:
    """Color codes in ANSI Escape code form for coloring terminal text."""

    BLUE = '\033[94m'
    CYAN = '\033[96m'
    GREEN = '\033[92m'
    LIME = '\033[38;5;190m'
    RED = '\033[91m'
    YELLOW = '\033[93m'
    WHITE = '\033[97m'
    ORANGE = '\033[38;5;214m'
    TURQ = '\033[36m'
    RESET = '\033[0m'


def convert_to_csv(excel_data, output_dir, base_name):
    """Convert Excel spreadsheet into CSV(Comma Seperated Values)."""
    try:
        csv_path = os.path.join(output_dir, f"{base_name}.csv")
        # Save to CSV with UTF-8 encoding with BOM
        excel_data.to_csv(csv_path, index=False, encoding='utf-8-sig')
    except Exception as e:
        print(
            Color.ORANGE + f"CSV dönüşümünde bir hata yaşandı: {e}"
            + Color.RESET
        )
    else:
        print(
            Color.YELLOW + "CSV dosyası başarıyla oluşturuldu."
            + Color.RESET
        )


def convert_to_json(excel_data, output_dir, base_name):
    """Convert Excel spreadsheet into JSON(JavaScript Object Notation)."""
    try:
        json_path = os.path.join(output_dir, f"{base_name}.json")
        json_data = excel_data.to_dict(orient='records')
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, ensure_ascii=False, indent=4)
    except Exception as e:
        print(
            Color.ORANGE + f"JSON dönüşümünde bir hata yaşandı: {e}"
            + Color.RESET
        )
    else:
        print(
            Color.YELLOW + "JSON dosyası başarıyla oluşturuldu."
            + Color.RESET
        )


def convert_to_xml(excel_data, output_dir, base_name):
    """Convert Excel spreadsheet into XML(Extensible Markup Language)."""
    try:
        xml_path = os.path.join(output_dir, f"{base_name}.xml")
        root = elementTree.Element("Table")

        def escape_xml(text):
            text = str(text)
            replacements = {
                '&': '&amp;',
                '<': '&lt;',
                '>': '&gt;',
                '"': '&quot;',
                "'": '&apos;'
            }
            for key, value in replacements.items():
                text = text.replace(key, value)
            return text

        for index, row in excel_data.iterrows():
            row_elem = elementTree.SubElement(root, "Row")
            for col_name in excel_data.columns:
                col_elem = elementTree.SubElement(row_elem, "Cell",
                                                  name=col_name)
                value = escape_xml(row[col_name])
                col_elem.text = value

        tree = elementTree.ElementTree(root)
        tree.write(xml_path, encoding='utf-8', xml_declaration=True)
    except Exception as e:
        print(
            Color.ORANGE + f"XML dönüşümünde bir hata yaşandı: {e}"
            + Color.RESET
        )
    else:
        print(
            Color.YELLOW + "XML dosyası başarıyla oluşturuldu."
            + Color.RESET
        )


def convert_to_html(excel_data, output_dir, base_name):
    """Convert Excel spreadsheet into HTML(HyperText Markup Language)."""
    try:
        html_path = os.path.join(output_dir, f"{base_name}.html")
        excel_data.to_html(html_path, index=False)
    except Exception as e:
        print(
            Color.ORANGE + f"HTML dönüşümünde bir hata yaşandı: {e}"
            + Color.RESET
        )
    else:
        print(
            Color.YELLOW + "HTML dosyası başarıyla oluşturuldu."
            + Color.RESET
        )


def convert_to_pdf(file_path, output_dir, base_name):
    """Convert Excel spreadsheet into PDF(Portable Document Format)."""
    try:
        print(
            Color.YELLOW + "PDF dönüşümü birkaç saniye alabilir... "
            + Color.RESET
        )

        pdf_path = os.path.join(output_dir, f"{base_name}.pdf")
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        pdf_doc = SimpleDocTemplate(
            pdf_path,
            pagesize=landscape(A4),
            rightMargin=10,
            leftMargin=10,
            topMargin=10,
            bottomMargin=10,
            title="TÜRASAŞ Demirbaş"
        )
        elements = []

        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(
                [str(cell) if cell is not None else '' for cell in row]
            )

        styles = getSampleStyleSheet()
        style = styles['BodyText']
        style.fontName = 'DejaVu'
        style.fontSize = 10
        style.wordWrap = 'CJK'
        style.alignment = 1

        page_width = landscape(A4)[0] - 20
        num_cols = len(data[0]) if data else 1
        col_width = page_width / num_cols

        wrapped_data = [
            [Paragraph(cell, style) for cell in row] for row in data
        ]

        table_style = TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'DejaVu'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('WORDWRAP', (0, 0), (-1, -1), 'CJK'),
        ])

        table = Table(wrapped_data, colWidths=[col_width] * num_cols)
        table.setStyle(table_style)
        elements.append(table)
        pdf_doc.build(elements)
    except Exception as e:
        print(
            Color.ORANGE + f"PDF dönüşümünde bir hata yaşandı: {e}"
            + Color.RESET
        )
    else:
        print(
            Color.YELLOW + "PDF dosyası başarıyla oluşturuldu."
            + Color.RESET
        )


def convert_excel(file_path, output_dir):
    """Handle all conversions written in the other functions."""
    try:
        try:
            excel_data = pd.read_excel(file_path)
            excel_data.fillna('', inplace=True)
            excel_data = excel_data.apply(lambda x: x.astype(str))
        except Exception as e:
            print(
                Color.YELLOW + f"Dosya okunmasında bir hata yaşandı: {e}"
                + Color.RESET
            )
            return
        else:
            print(
                Color.YELLOW + "Excel dosyası başarıyla okundu." + Color.RESET
            )

        base_name = os.path.splitext(os.path.basename(file_path))[0]

        # Create threads for each conversion function
        threads = [
            threading.Thread(target=convert_to_csv, args=(excel_data, output_dir, base_name)),
            threading.Thread(target=convert_to_json, args=(excel_data, output_dir, base_name)),
            threading.Thread(target=convert_to_xml, args=(excel_data, output_dir, base_name)),
            threading.Thread(target=convert_to_html, args=(excel_data, output_dir, base_name)),
            threading.Thread(target=convert_to_pdf, args=(file_path, output_dir, base_name))
        ]

        # Start all the threads
        for thread in threads:
            thread.start()

        # Wait for all threads to complete
        for thread in threads:
            thread.join()

    except Exception as e:
        print(
            Color.ORANGE + f"Beklenmeyen bir hata yaşandı: {e}"
            + Color.RESET
        )
    else:
        print(
            Color.LIME + "Tüm dönüşümler başarıyla tamamlandı!"
            + Color.RESET
        )


def handle_file_selection():
    """Handle the file selection interface."""
    root = tk.Tk()
    root.title("File Selection")
    root.geometry("0x0+0+0")
    root.update()
    root.update_idletasks()
    root.attributes('-topmost', False)
    root.attributes('-topmost', True)

    file_path = filedialog.askopenfilename(
        title="Bir Excel dosyası seçiniz",
        filetypes=[("Excel files", "*.xls *.xlsx *.xlsm *.xlsb")])

    if not file_path:
        print(Color.ORANGE + "Dosya seçilemedi, lütfen tekrar deneyiniz")
        root.destroy()
        return

    if not file_path.endswith((".xls", ".xlsx", ".xlsm", ".xlsb")):
        messagebox.showerror(
            "Hata",
            "Seçilen dosya bir Excel dosyası değildir.")
        root.destroy()
        return

    output_dir = filedialog.askdirectory(title="Çıktı Dizinini Seçiniz")
    if not output_dir:
        messagebox.showerror("Hata", "Lütfen geçerli bir dizin seçiniz.")
        root.destroy()
        return

    base_name = os.path.splitext(os.path.basename(file_path))[0]
    output_dir = os.path.join(output_dir, base_name)

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    convert_excel(file_path, output_dir)

    messagebox.showinfo(
        "Başarılı!", f"Dosya başarıyla dönüştürüldü. Dizin:\n {output_dir}")
    root.destroy()


def main():
    """Main"""
    os.system('cls')

    welcome_string = (
    f"{Color.RESET}Hoş geldiniz. Bu program Excel ile uyumlu "
    "(xls, xlsx, xlsm, xlsb) dosyaları "
    "json, csv, html, xml ve pdf "
    "dosyalarına dönüştürür.")
    print()
    print(Color.TURQ + ascii_banner + Color.RESET)
    print(welcome_string)

    while True:
        print(Color.RESET + "Dosya seçmek için ", end='')
        print(Color.GREEN + "1" + Color.RESET, end='')
        print(", çıkış yapmak için ", end='')
        print(Color.RED + "2" + Color.RESET, end='')
        print(" tuşuna basınız:")

        input_value = input()

        if input_value == "1":
            handle_file_selection()
        elif input_value == "2":
            print(Color.YELLOW + "Çıkış yapılıyor..." + Color.RESET)
            temp_file_path = 'temp.csv'
            if os.path.exists(temp_file_path): os.remove(temp_file_path)
            time.sleep(1)
            os.system('cls')
            sys.exit(0)
        else:
            print(Color.YELLOW
                  + "Hatalı giriş yaptınız, lütfen tekrar deneyiniz."
                  + Color.RESET)


# Make the program work both as a standalone script and an importable module
if __name__ == "__main__":
    main()
